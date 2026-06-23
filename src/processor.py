import pandas as pd
from typing import List, Dict, Optional
import os


class DataProcessor:
    """Handles Excel data extraction and filtering for active sales items.
    
    Implements business report generation with automatic column detection,
    compact transaction details, and comprehensive summary sections.
    """
    
    def __init__(self):
        self.data: Optional[pd.DataFrame] = None
        self.filtered_data: Optional[pd.DataFrame] = None
        self.is_real_time_format = False
        self.business_columns: List[str] = []
    
    def _is_real_time_file(self, file_path: str) -> bool:
        """Check if the file is a Real Time Customer Order Payment List Report."""
        filename = os.path.basename(file_path)
        return filename.startswith("Real Time")
    
    def load_excel(self, file_path: str) -> bool:
        """Load data from Excel file and auto-detect business columns."""
        try:
            self.is_real_time_format = self._is_real_time_file(file_path)
            
            if self.is_real_time_format:
                # Real Time format has header at row 10 (0-indexed)
                self.data = pd.read_excel(file_path, header=10)
            else:
                # Standard format
                self.data = pd.read_excel(file_path)
            
            # Auto-detect business-relevant columns
            self._detect_business_columns()
            
            return True
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return False
    
    def _detect_business_columns(self):
        """Automatically identify business-relevant columns from the dataset.
        
        Ignores technical fields, metadata, audit fields, empty columns, 
        duplicate columns, and unnecessary information.
        """
        if self.data is None:
            return
        
        # Common business column patterns to look for
        business_patterns = [
            'date', 'time', 'customer', 'name', 'amount', 'price', 'cost',
            'quantity', 'total', 'status', 'category', 'type', 'ref', 'reference',
            'order', 'invoice', 'number', 'service', 'operation', 'payment'
        ]
        
        # Technical/metadata patterns to ignore
        ignore_patterns = [
            'id', 'uuid', 'guid', 'timestamp', 'created', 'updated', 'modified',
            'audit', 'meta', 'system', 'internal', 'temp', 'flag', 'active'
        ]
        
        self.business_columns = []
        
        for column in self.data.columns:
            col_lower = str(column).lower()
            
            # Skip empty columns
            if self.data[column].isna().all():
                continue
            
            # Skip technical/metadata columns
            if any(pattern in col_lower for pattern in ignore_patterns):
                continue
            
            # Include columns that match business patterns
            if any(pattern in col_lower for pattern in business_patterns):
                self.business_columns.append(column)
        
        # If no business columns found, use all non-empty columns
        if not self.business_columns:
            self.business_columns = [
                col for col in self.data.columns 
                if not self.data[col].isna().all()
            ]
    
    def filter_active_sales(self, 
                           status_column: str = "status",
                           active_value: str = "active") -> pd.DataFrame:
        """Filter rows where status is active/has sales."""
        if self.data is None:
            raise ValueError("No data loaded. Call load_excel first.")
        
        # Use appropriate column name based on file format
        if self.is_real_time_format:
            status_column = "Order Status"
            active_value = "Completed"  # Default for Real Time files
        
        # Filter for active items (case-insensitive)
        if status_column in self.data.columns:
            mask = self.data[status_column].str.lower() == active_value.lower()
            self.filtered_data = self.data[mask].copy()
        else:
            # If column doesn't exist, return all data
            print(f"Warning: Column '{status_column}' not found. Returning all data.")
            self.filtered_data = self.data.copy()
        
        return self.filtered_data
    
    def filter_specific_transactions(self) -> pd.DataFrame:
        """Filter for specific transaction types: 85.00Birr, 100.00Birr, and zero price operations."""
        if self.data is None:
            raise ValueError("No data loaded. Call load_excel first.")
        
        if not self.is_real_time_format:
            print("Warning: Specific transaction filtering is only available for Real Time format.")
            return self.data.copy()
        
        # First filter by Order Status = Completed
        if "Order Status" in self.data.columns:
            status_mask = self.data["Order Status"].str.lower() == "completed"
            filtered_by_status = self.data[status_mask].copy()
        else:
            filtered_by_status = self.data.copy()
        
        # Clean payment amounts for filtering
        filtered_by_status["CleanAmount"] = filtered_by_status["Total Payment Amount"].apply(
            lambda x: self._extract_numeric_amount(x)
        )
        
        # Filter for specific amounts: 85.00, 100.00, and 0.00
        amount_mask = (
            (filtered_by_status["CleanAmount"] == 85.00) |
            (filtered_by_status["CleanAmount"] == 100.00) |
            (filtered_by_status["CleanAmount"] == 0.00)
        )
        
        # For zero price transactions, also check Business Operation column
        zero_price_mask = filtered_by_status["CleanAmount"] == 0.00
        if "Business Operation" in filtered_by_status.columns:
            # Include zero price transactions with specific business operations
            business_operations_to_include = [
                "Change Subscriber SIM Card",
                "Create Customer", 
                "Change Supplementary Offering"
            ]
            business_op_mask = filtered_by_status["Business Operation"].isin(business_operations_to_include)
            # Combine: either (zero price AND valid business op) OR (85/100 birr)
            final_mask = (
                (zero_price_mask & business_op_mask) |
                (filtered_by_status["CleanAmount"] == 85.00) |
                (filtered_by_status["CleanAmount"] == 100.00)
            )
        else:
            # If no Business Operation column, just use amount filter
            final_mask = amount_mask
        
        self.filtered_data = filtered_by_status[final_mask].copy()
        
        # Add categorization
        self.filtered_data["Category"] = self.filtered_data.apply(self._categorize_transaction, axis=1)
        
        # Remove the temporary CleanAmount column
        if "CleanAmount" in self.filtered_data.columns:
            del self.filtered_data["CleanAmount"]
        
        return self.filtered_data
    
    def _categorize_transaction(self, row) -> str:
        """Categorize transaction based on business operation and amount."""
        # Get clean amount
        amount_str = str(row.get('Total Payment Amount', ''))
        amount = self._extract_numeric_amount(amount_str)
        
        business_op = str(row.get('Business Operation', '')).lower()
        
        # Categorization logic
        if amount == 85.0:
            return 'Offer'  # 85 birr transactions
        elif amount == 100.0 and 'change subscriber sim card' in business_op:
            return 'Replacement'  # 100 birr SIM card changes
        elif amount == 0.0:
            return 'Transfer'  # Zero price transactions
        elif amount == 100.0:
            return 'Bundle'  # Other 100 birr transactions (potential bundles)
        
        return 'Other'
    
    def _extract_numeric_amount(self, amount):
        """Extract numeric value from payment amount string."""
        if pd.isna(amount) or amount == "":
            return 0.0
        
        amount_str = str(amount)
        # Remove 'Birr' suffix, commas, and whitespace
        amount_str = amount_str.replace("Birr", "").replace(",", "").strip()
        
        try:
            return float(amount_str)
        except (ValueError, TypeError):
            return 0.0
    
    def format_for_printing(self, 
                           columns_to_include: Optional[List[str]] = None) -> pd.DataFrame:
        """Format data for A5 paper printing."""
        if self.filtered_data is None:
            raise ValueError("No filtered data available.")
        
        if columns_to_include:
            # Select only specified columns if they exist
            available_cols = [col for col in columns_to_include if col in self.filtered_data.columns]
            if available_cols:
                formatted_data = self.filtered_data[available_cols].copy()
            else:
                formatted_data = self.filtered_data.copy()
        else:
            formatted_data = self.filtered_data.copy()
        
        # Clean and format data
        formatted_data = formatted_data.fillna("")
        
        # Clean payment amounts if in Real Time format
        if self.is_real_time_format and "Total Payment Amount" in formatted_data.columns:
            formatted_data["Total Payment Amount"] = formatted_data["Total Payment Amount"].apply(
                lambda x: self._clean_payment_amount(x)
            )
        
        return formatted_data
    
    def _clean_payment_amount(self, amount):
        """Clean payment amount by removing 'Birr' suffix and converting to proper format."""
        if pd.isna(amount) or amount == "":
            return ""
        
        amount_str = str(amount)
        # Remove 'Birr' suffix, commas, and any whitespace
        amount_str = amount_str.replace("Birr", "").replace(",", "").strip()
        
        # Format to 2 decimal places if it's a number
        try:
            num_amount = float(amount_str)
            return f"{num_amount:.2f}"
        except (ValueError, TypeError):
            return amount_str
    
    def get_summary(self) -> Dict:
        """Get summary statistics of the filtered data."""
        if self.filtered_data is None:
            return {"error": "No filtered data available"}
        
        return {
            "total_items": len(self.filtered_data),
            "columns": list(self.filtered_data.columns),
            "preview": self.filtered_data.head(5).to_dict()
        }
    
    def get_category_summary(self) -> Dict:
        """Get summary statistics by category."""
        if self.filtered_data is None or "Category" not in self.filtered_data.columns:
            return {"error": "No categorized data available"}
        
        category_summary = {}
        
        for category in ['Offer', 'Replacement', 'Transfer', 'Bundle']:
            cat_data = self.filtered_data[self.filtered_data['Category'] == category]
            if len(cat_data) > 0:
                # Calculate total amount for this category
                total_amount = cat_data['Total Payment Amount'].apply(
                    lambda x: self._extract_numeric_amount(x)
                ).sum()
                
                category_summary[category] = {
                    "count": len(cat_data),
                    "total_amount": total_amount
                }
        
        return category_summary
    
    def add_summary_to_excel(self, data: pd.DataFrame, output_path: str) -> bool:
        """Export data with category summary to Excel file."""
        try:
            if data is None:
                raise ValueError("No data to export")
            
            # Calculate category summary
            category_summary = {}
            for category in ['Offer', 'Replacement', 'Transfer', 'Bundle']:
                cat_data = data[data['Category'] == category]
                if len(cat_data) > 0:
                    total_amount = cat_data['Total Payment Amount'].apply(
                        lambda x: self._extract_numeric_amount(x)
                    ).sum()
                    category_summary[category] = {
                        "count": len(cat_data),
                        "total_amount": total_amount
                    }
            
            # Create summary DataFrame with proper column names
            summary_data = []
            for category, info in category_summary.items():
                summary_data.append([category, info["count"], f"{info['total_amount']:.2f}"])
            
            # Combine data and summary using openpyxl directly
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            
            # First write the main data
            data.to_excel(output_path, index=False, engine='openpyxl')
            
            # Then load the workbook and add the summary
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Find the next empty row after the data
            start_row = len(data) + 4  # Leave 3 empty rows for spacing
            
            # Clear cells in the summary section to avoid column confusion
            for row in range(start_row, start_row + 10):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col, value=None)
            
            # Write summary header
            ws.cell(row=start_row, column=1, value="CATEGORY SUMMARY")
            ws.cell(row=start_row, column=1).font = Font(bold=True)
            
            # Write summary column headers
            ws.cell(row=start_row + 1, column=1, value="Category")
            ws.cell(row=start_row + 1, column=2, value="Count")
            ws.cell(row=start_row + 1, column=3, value="Total Amount (Birr)")
            
            # Write summary data
            for i, (category, count, amount) in enumerate(summary_data):
                ws.cell(row=start_row + 2 + i, column=1, value=category)
                ws.cell(row=start_row + 2 + i, column=2, value=count)
                ws.cell(row=start_row + 2 + i, column=3, value=amount)
            
            wb.save(output_path)
            
            return True
        except Exception as e:
            print(f"Error exporting to Excel with summary: {e}")
            return False
    
    def export_to_excel(self, output_path: str, data: Optional[pd.DataFrame] = None, with_summary: bool = False) -> bool:
        """Export filtered data to Excel file."""
        try:
            export_data = data if data is not None else self.filtered_data
            if export_data is None:
                raise ValueError("No data to export")
            
            # Use summary export if data has categories and with_summary is True
            if with_summary and "Category" in export_data.columns:
                return self.add_summary_to_excel(export_data, output_path)
            else:
                export_data.to_excel(output_path, index=False)
                return True
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def format_for_report(self) -> pd.DataFrame:
        """Format filtered data for the specific report structure: ID1, ID2, Transaction, Date, User, Reference, Branch."""
        if self.filtered_data is None:
            raise ValueError("No filtered data available for formatting.")
        
        # Map existing columns to the required structure
        formatted_data = pd.DataFrame()
        
        # Map columns based on available data
        if "Service Number" in self.filtered_data.columns:
            formatted_data["ID1"] = self.filtered_data["Service Number"]
        elif "Customer ID" in self.filtered_data.columns:
            formatted_data["ID1"] = self.filtered_data["Customer ID"]
        else:
            formatted_data["ID1"] = ""
        
        if "Order No" in self.filtered_data.columns:
            formatted_data["ID2"] = self.filtered_data["Order No"]
        elif "Order Number" in self.filtered_data.columns:
            formatted_data["ID2"] = self.filtered_data["Order Number"]
        else:
            formatted_data["ID2"] = ""
        
        # Format Transaction column: combine Business Operation and Amount
        if "Business Operation" in self.filtered_data.columns and "Total Payment Amount" in self.filtered_data.columns:
            formatted_data["Transaction"] = self.filtered_data.apply(
                lambda row: f"{row['Business Operation']} at {self._clean_payment_amount(row['Total Payment Amount'])}.",
                axis=1
            )
        elif "Business Operation" in self.filtered_data.columns:
            formatted_data["Transaction"] = self.filtered_data["Business Operation"]
        else:
            formatted_data["Transaction"] = "Payment"
        
        # Map Date column
        if "Date" in self.filtered_data.columns:
            formatted_data["Date"] = pd.to_datetime(self.filtered_data["Date"], dayfirst=True).dt.strftime("%d-%m-%Y")
        elif "Sales Date and Time" in self.filtered_data.columns:
            formatted_data["Date"] = pd.to_datetime(self.filtered_data["Sales Date and Time"], dayfirst=True).dt.strftime("%d-%m-%Y")
        else:
            formatted_data["Date"] = ""
        
        # Map User column
        if "Customer Name" in self.filtered_data.columns:
            formatted_data["User"] = self.filtered_data["Customer Name"]
        elif "User" in self.filtered_data.columns:
            formatted_data["User"] = self.filtered_data["User"]
        else:
            formatted_data["User"] = ""
        
        # Map Reference column - try to find a better reference field
        if "Reference" in self.filtered_data.columns:
            formatted_data["Reference"] = self.filtered_data["Reference"]
        elif "Ref" in self.filtered_data.columns:
            formatted_data["Reference"] = self.filtered_data["Ref"]
        elif "Service Number" in self.filtered_data.columns:
            formatted_data["Reference"] = self.filtered_data["Service Number"]
        else:
            formatted_data["Reference"] = ""
        
        # Map Branch column
        if "Branch" in self.filtered_data.columns:
            formatted_data["Branch"] = self.filtered_data["Branch"]
        elif "Location" in self.filtered_data.columns:
            formatted_data["Branch"] = self.filtered_data["Location"]
        elif "Site" in self.filtered_data.columns:
            formatted_data["Branch"] = self.filtered_data["Site"]
        else:
            formatted_data["Branch"] = ""
        
        # Fill empty values
        formatted_data = formatted_data.fillna("")
        
        return formatted_data
    
    def generate_business_report(self) -> str:
        """Generate a business report with the specific table structure and summary format.
        
        Returns a formatted string containing:
        - Table with columns: ID1, ID2, Transaction, Date, User, Reference, Branch (limited rows)
        - Summary section with category grouping (without "Summary" header)
        - Category counts and amounts
        - Grand total
        """
        if self.filtered_data is None:
            return "No filtered data available for report generation."
        
        # Format data for the specific structure
        report_data = self.format_for_report()
        
        # Generate the report
        report_lines = []
        
        # Create table header
        separator = "+" + "-" * 10 + "+" + "-" * 11 + "+" + "-" * 22 + "+" + "-" * 12 + "+" + "-" * 10 + "+" + "-" * 12 + "+" + "-" * 12 + "+"
        report_lines.append(separator)
        
        # Add column headers
        headers = "| ID1      | ID2       | Transaction           | Date       | User     | Reference  | Branch     |"
        report_lines.append(headers)
        report_lines.append(separator)
        
        # Add data rows (limited to first 10 rows)
        max_rows = 10
        for idx, row in report_data.head(max_rows).iterrows():
            id1 = str(row["ID1"])[:10].ljust(10)
            id2 = str(row["ID2"])[:11].ljust(11)
            transaction = str(row["Transaction"])[:22].ljust(22)  # Truncate to 22 chars
            date = str(row["Date"])[:12].ljust(12)
            user = str(row["User"])[:10].ljust(10)
            reference = str(row["Reference"])[:12].ljust(12)
            branch = str(row["Branch"])[:12].ljust(12)
            
            row_str = f"| {id1} | {id2} | {transaction} | {date} | {user} | {reference} | {branch} |"
            report_lines.append(row_str)
        
        report_lines.append(separator)
        report_lines.append("")
        
        # Generate summary section (without "Summary" header)
        # Determine category column
        category_col = None
        potential_cat_cols = ['Category', 'category', 'Type', 'type', 'Business Operation', 'business operation']
        for col in potential_cat_cols:
            if col in self.filtered_data.columns:
                category_col = col
                break
        
        # Determine amount column
        amount_col = None
        potential_amt_cols = ['Total Payment Amount', 'total payment amount', 'Amount', 'amount', 'Price', 'price']
        for col in potential_amt_cols:
            if col in self.filtered_data.columns:
                amount_col = col
                break
        
        if category_col and amount_col:
            # Group by category and calculate statistics
            category_groups = self.filtered_data.groupby(category_col)
            
            grand_total = 0
            for category, group in category_groups:
                count = len(group)
                total_amount = group[amount_col].apply(
                    lambda x: self._extract_numeric_amount(x)
                ).sum()
                grand_total += total_amount
                # Format: category name (left), count (right), amount (right)
                report_lines.append(f"{str(category):<15} {count:>6} {total_amount:>10.2f}")
            
            report_lines.append("-" * 35)
            report_lines.append(f"{'total':<15} {'':>6} {grand_total:>10.2f}")
        else:
            # Fallback: basic statistics
            report_lines.append(f"Total Records: {len(report_data)}")
            if amount_col:
                total_amount = self.filtered_data[amount_col].apply(
                    lambda x: self._extract_numeric_amount(x)
                ).sum()
                report_lines.append(f"Total Amount: {total_amount:.2f}")
        
        return "\n".join(report_lines)